import openpyxl
import os
import sys

path = r"C:\Users\mmk_b\Downloads"
files = [f for f in os.listdir(path) if f.startswith("PTY") and f.endswith(".xlsx")]
files.sort(key=lambda x: os.path.getmtime(os.path.join(path, x)), reverse=True)

output = []

if not files:
    output.append("No PTY files found")
else:
    target_file = os.path.join(path, files[0])
    output.append(f"Reading file: {target_file}")
    wb = openpyxl.load_workbook(target_file, data_only=True)
    sheet = wb.active
    output.append(f"Active Sheet Name: {sheet.title}")
    
    # Read row 1
    row1 = [cell for cell in next(sheet.iter_rows(max_row=1, values_only=True))]
    output.append("Row 1 cells:")
    for idx, val in enumerate(row1):
        col_letter = openpyxl.utils.get_column_letter(idx + 1)
        output.append(f"{col_letter}: {val}")
        
    output.append("\nFirst 10 data rows (cols A to T):")
    row_count = 0
    for row in sheet.iter_rows(min_row=2, max_row=15, values_only=True):
        row_count += 1
        output.append(f"Row {row_count+1}: {[row[i] for i in [0,1,2,3,8,19] if i < len(row)]}") # indices for A, B, C, D, I, T (0, 1, 2, 3, 8, 19)

with open('excel_inspect_results.txt', 'w', encoding='utf-8') as f:
    f.write("\n".join(output))
print("Done! Saved to excel_inspect_results.txt")
